# -*- coding: utf-8 -*-
import openpyxl
from append_xlsx import append_rows

HEADER = ["Bài", "生词", "Pinyin", "描述", "意义", "例如", "复习", "检查"]


def _make_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Từ vựng"
    ws.append(HEADER)
    ws.append(["Bài 1", "已经", "yǐjīng", "", "đã", "我已经吃了。", "", ""])
    wb.save(str(path))


def test_append_new_row(tmp_path):
    xlsx = tmp_path / "tv.xlsx"
    _make_xlsx(xlsx)
    payload = {"bai": 5, "words": [
        {"w": "尴尬", "pinyin": "gāngà", "desc": "不好意思", "vi": "bối rối",
         "ex": "很尴尬。"}]}
    n = append_rows(payload, xlsx)
    assert n == 1
    wb = openpyxl.load_workbook(str(xlsx))
    ws = wb["Từ vựng"]
    rows = list(ws.iter_rows(values_only=True))
    last = rows[-1]
    assert last[0] == "Bài 5" and last[1] == "尴尬" and last[2] == "gāngà"
    assert last[4] == "bối rối"


def test_dedup_existing(tmp_path):
    xlsx = tmp_path / "tv.xlsx"
    _make_xlsx(xlsx)
    payload = {"bai": 2, "words": [
        {"w": "已经", "pinyin": "yǐjīng", "desc": "", "vi": "đã", "ex": ""}]}
    n = append_rows(payload, xlsx)
    assert n == 0  # 已经 đã có
