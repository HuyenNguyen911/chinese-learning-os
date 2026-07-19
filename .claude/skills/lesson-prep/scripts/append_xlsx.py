# -*- coding: utf-8 -*-
"""append_xlsx.py — append từ mới vào raw/Từ vựng.xlsx (sheet 'Từ vựng').

Dedup theo 生词. Cột: Bài, 生词, Pinyin, 描述, 意义, 例如, 复习, 检查.
Chạy:  python append_xlsx.py <vocab_payload.json> [xlsx]
"""
import sys, json
from pathlib import Path
import openpyxl


def _sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.lower():
            return wb[s]
    return wb[wb.sheetnames[0]]


def append_rows(payload, xlsx_path):
    xlsx_path = str(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = _sheet(wb, "Từ vựng")
    have = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) > 1 and row[1]:
            have.add(str(row[1]).strip())
    bai = payload.get("bai")
    added = 0
    for it in payload.get("words", []):
        w = (it.get("w") or "").strip()
        if not w or w in have:
            continue
        have.add(w)
        ws.append(["Bài %s" % bai if bai is not None else "",
                   w, it.get("pinyin", ""), it.get("desc", ""),
                   it.get("vi", ""), it.get("ex", ""), "", ""])
        added += 1
    if added:
        wb.save(xlsx_path)
    return added


def main(argv):
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8")
        except (AttributeError, ValueError):
            pass
    if len(argv) < 2:
        print("Usage: python append_xlsx.py <vocab_payload.json> [xlsx]",
              file=sys.stderr); return 2
    payload = json.loads(Path(argv[1]).read_text(encoding="utf-8"))
    xlsx = argv[2] if len(argv) > 2 else "raw/Từ vựng.xlsx"
    n = append_rows(payload, xlsx)
    print("append_xlsx: +%d từ -> %s (sheet 'Từ vựng')" % (n, xlsx))
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
