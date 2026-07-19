# -*- coding: utf-8 -*-
# Đọc raw/Từ vựng.xlsx (2 sheet) -> data/tv.json (生词) + data/ct.json (生词拓展)
#   sheet 'Từ vựng' : cột Bài, 生词, Pinyin, 描述, 意义, 例如, 复习, 检查
#   sheet 'Chung từ': cột Bài, 生词(=chuỗi nhóm họ từ), ..., 意义(ghi chú Việt)
# Chạy từ gốc repo:  python .claude/skills/vocab-study/scripts/extract_xlsx.py [đường_dẫn_xlsx]
import openpyxl, json, re, os, sys
for _s in (sys.stdout, sys.stderr):  # console Windows cp1252 → tránh crash khi print 中文/tiếng Việt
    try: _s.reconfigure(encoding="utf-8")
    except Exception: pass

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "..", "data")
XLSX = sys.argv[1] if len(sys.argv) > 1 else "raw/Từ vựng.xlsx"

def bnum(v):
    m = re.match(r"\s*[Bb]ài\s*(\d+)", str(v or ""))
    return int(m.group(1)) if m else None

def clean(s):
    return re.sub(r"[\x00-\x1f  ﻿​-‏]", " ", str(s or "")).strip()

wb = openpyxl.load_workbook(XLSX, data_only=True)

def sheet(name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.lower():
            return wb[s]
    return None

tv, ct = {}, {}
ws = sheet("Từ vựng") or wb[wb.sheetnames[0]]
for row in ws.iter_rows(min_row=2, values_only=True):
    n = bnum(row[0])
    if not n:
        continue
    w = clean(row[1])
    if not w:
        continue
    tv.setdefault(str(n), []).append({
        "w": w, "desc": clean(row[3]), "vi": clean(row[4]), "ex": clean(row[5]),
    })

ws2 = sheet("Chung từ")
if ws2:
    for row in ws2.iter_rows(min_row=2, values_only=True):
        n = bnum(row[0])
        if not n:
            continue
        grp = clean(row[1])
        if not grp:
            continue
        ct.setdefault(str(n), []).append({"grp": grp, "vi": clean(row[4])})

json.dump(tv, open(os.path.join(DATA, "tv.json"), "w", encoding="utf-8"), ensure_ascii=False)
json.dump(ct, open(os.path.join(DATA, "ct.json"), "w", encoding="utf-8"), ensure_ascii=False)
print("extract OK | 生词 bài:", len(tv), "tổng từ:", sum(len(v) for v in tv.values()),
      "| 拓展 bài:", len(ct))
